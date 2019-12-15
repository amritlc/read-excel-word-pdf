import openpyxl, os

os.getcwd() # check working directory
# os.chdir('c:\\Users\\dell\\Desktop\\AutomateWithPython') => change directory

workbook = openpyxl.load_workbook('example.xlsx')

print(type(workbook))

sheet = workbook.get_sheet_by_name('Sheet1')
print(type(sheet))

print(workbook.get_sheet_names())

print(sheet['A1'])

cell = sheet['A1']
print(cell.value)

A1 = str(sheet['A1'].value)
print(A1)

C1 = sheet['C1'].value
print(C1)

cell = sheet.cell(row=1,column=2)
print(cell)


for i in range(1,8):
	print(i, sheet.cell(row=i,column=2).value)




